#!/usr/bin/env python3
"""testbook-export, executed for real on US-EVAL-009 (one-off session script, D42).

Deliverables produced (SKILL.md "Deliverables (D25)"):
  1. .feature copied as-is
  2. synthesis.md copied as-is (re-projection = copy; this run does not redefine it)
  3. XLSX workbook, 3 sheets (Scenarios / Coverage matrix / Decisions & assumptions)
  4. opt-in Xray CSV  (connectors/xray.md mapping table)
  4b. opt-in TestRail CSV (connectors/testrail.md mapping table)

Every CSV is re-parsed after writing (connectors step 4: "CSV writing + re-parse check").
Run: python eval/skill-coverage-wave-2026-07-30/A-artifacts/testbook-export/export.py
"""
import csv, io, os, re, shutil, sys

ROOT = os.getcwd()
SRC = os.path.join(ROOT, "eval", "skill-eval-campaign-2026-07-29", "US-EVAL-009-octoperf-petstore")
# SKILL.md step 1 default target location: .qaia/testbooks/<US-ID>/export/
OUT = os.path.join(ROOT, "eval", "skill-coverage-wave-2026-07-30", "A-artifacts",
                   "testbook-export", "export")
FEATURE = os.path.join(SRC, "testbooks", "octoperf-petstore-cart.feature")
MATRIX = os.path.join(SRC, "testbooks", "coverage-matrix.md")
US = "US-EVAL-009"

TECH_TAGS = {"@ep", "@boundary", "@decision-table", "@state-transition", "@use-case",
             "@pairwise", "@error-guessing", "@checklist"}
STEP_KW = ("Given ", "When ", "Then ", "And ", "But ", "* ")


def read(p):
    return open(p, encoding="utf-8").read()


# ---------------------------------------------------------------- parse .feature
text = read(FEATURE)
lines = text.splitlines()

feature_title = next(l.split(":", 1)[1].strip() for l in lines if l.strip().startswith("Feature:"))

# Background steps (flattened into every row per connectors/xray.md; none in this book)
background = []
in_bg = False
for l in lines:
    s = l.strip()
    if s.startswith("Background:"):
        in_bg = True
        continue
    if in_bg:
        if s.startswith("Scenario") or s.startswith("@"):
            in_bg = False
        elif s.startswith(STEP_KW):
            background.append(s)

scenarios = []
cur = None
pending_tags, pending_comments = [], []
for l in lines:
    s = l.strip()
    if s.startswith("@"):
        if cur:
            scenarios.append(cur); cur = None
        pending_tags = s.split()
        pending_comments = []
    elif s.startswith("#") and pending_tags:
        pending_comments.append(s.lstrip("# ").rstrip())
    elif s.startswith("Scenario Outline:") or s.startswith("Scenario:"):
        if cur:
            scenarios.append(cur)
        cur = {"tags": pending_tags, "comments": list(pending_comments),
               "outline": s.startswith("Scenario Outline:"),
               "keyword": "Scenario Outline" if s.startswith("Scenario Outline:") else "Scenario",
               "title": s.split(":", 1)[1].strip(), "steps": [], "examples": []}
        pending_tags, pending_comments = [], []
    elif cur is not None and s.startswith(STEP_KW):
        cur["steps"].append(s)
    elif cur is not None and s.startswith("Examples:"):
        cur["in_examples"] = True
    elif cur is not None and cur.get("in_examples") and s.startswith("|"):
        cur["examples"].append(s)
if cur:
    scenarios.append(cur)

if not scenarios:
    sys.exit("BLOCKER: no scenario parsed from the .feature")


def tag(sc, pred, default=""):
    for t in sc["tags"]:
        if pred(t):
            return t
    return default


def enrich(sc):
    sc["id"] = tag(sc, lambda t: t.startswith("@QAIA-"))[1:]
    sc["ac"] = tag(sc, lambda t: re.fullmatch(r"@AC\d+", t))[1:]
    sc["priority"] = tag(sc, lambda t: t in ("@P1", "@P2", "@P3"))[1:]
    sc["technique"] = ", ".join(t[1:] for t in sc["tags"] if t in TECH_TAGS)
    sc["oracle"] = ", ".join(t.split(":", 1)[1] for t in sc["tags"] if t.startswith("@oracle:"))
    sc["negative"] = "yes" if "@negative" in sc["tags"] else "no"
    sc["confidence"] = "low" if "@low-confidence" in sc["tags"] else "full"
    m = re.search(r"condition:\s*(AC\d+-C\d+)", " ".join(sc["comments"]))
    sc["condition"] = m.group(1) if m else ""
    rationale = " ".join(c for c in sc["comments"] if not c.startswith("condition:"))
    if sc["confidence"] == "low":
        rationale = (rationale + " [low confidence]").strip()
    sc["rationale"] = re.sub(r"\s+", " ", rationale).strip()
    return sc


scenarios = [enrich(s) for s in scenarios]


def rows(sc):
    """Explode a Scenario Outline into one row per Examples line, ID suffixed -eN."""
    body = [r for r in sc["examples"][1:]] if sc["outline"] else []
    if not body:
        return [(sc["id"], sc["title"], sc["steps"])]
    out = []
    for n, _ in enumerate(body, 1):
        out.append((f"{sc['id']}-e{n}", f"{sc['title']} - Examples {n}", sc["steps"]))
    return out


# ---------------------------------------------------------------- 1 & 2: copies
os.makedirs(OUT, exist_ok=True)
shutil.copy2(FEATURE, os.path.join(OUT, os.path.basename(FEATURE)))
shutil.copy2(os.path.join(SRC, "testbooks", "synthesis.md"), os.path.join(OUT, "synthesis.md"))

# ---------------------------------------------------------------- 4: Xray CSV
XRAY_PRI = {"P1": "Highest", "P2": "Medium", "P3": "Low"}
xray_dir = os.path.join(OUT, "xray"); os.makedirs(xray_dir, exist_ok=True)
xray_path = os.path.join(xray_dir, f"{US}-xray.csv")
xray_header = ["Issue Type", "Test Type", "Cucumber Test Type", "Summary",
               "Cucumber Scenario", "Labels", "Priority", "Description"]
with open(xray_path, "w", encoding="utf-8", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(xray_header)
    for sc in scenarios:
        for rid, rtitle, steps in rows(sc):
            labels = " ".join([rid, sc["priority"]] +
                              [t[1:] for t in sc["tags"] if t in TECH_TAGS])
            w.writerow(["Test", "Automated[Cucumber]", sc["keyword"], rtitle,
                        "\n".join(background + steps), labels,
                        XRAY_PRI.get(sc["priority"], ""),
                        f"condition: {sc['condition']}. {sc['rationale']}".strip()])

# ---------------------------------------------------------------- 4b: TestRail CSV
TR_PRI = {"P1": "Critical", "P2": "Medium", "P3": "Low"}
tr_dir = os.path.join(OUT, "testrail"); os.makedirs(tr_dir, exist_ok=True)
tr_path = os.path.join(tr_dir, f"{US}-testrail.csv")
tr_header = ["Title", "Type", "Priority", "Preconditions", "Steps",
             "Expected Result", "Section", "References"]
empty_steps = []
with open(tr_path, "w", encoding="utf-8", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(tr_header)
    for sc in scenarios:
        for rid, rtitle, steps in rows(sc):
            i_when = next((i for i, s in enumerate(steps) if s.startswith("When ")), None)
            i_then = next((i for i, s in enumerate(steps) if s.startswith("Then ")), None)
            if i_when is None:
                pre, act = steps[:i_then] if i_then is not None else steps, []
            else:
                pre, act = steps[:i_when], steps[i_when:i_then]
            exp = steps[i_then:] if i_then is not None else []
            if not act:
                empty_steps.append(rid)
            refs = f"{rid} {sc['priority']} {sc['technique']} | condition: {sc['condition']} | {sc['rationale']}"
            w.writerow([rtitle, "Automated", TR_PRI.get(sc["priority"], ""),
                        "\n".join(background + pre), "\n".join(act), "\n".join(exp),
                        feature_title, refs.strip()])

# ---------------------------------------------------------------- re-parse check
report = []
for label, path, header in (("xray", xray_path, xray_header), ("testrail", tr_path, tr_header)):
    with open(path, encoding="utf-8", newline="") as fh:
        data = list(csv.reader(fh))
    assert data[0] == header, f"{label}: header drift {data[0]}"
    assert all(len(r) == len(header) for r in data), f"{label}: ragged row"
    report.append(f"re-parse OK  {label:9} rows={len(data)-1} cols={len(header)} -> {path}")

# ---------------------------------------------------------------- 3: XLSX
try:
    from openpyxl import Workbook
except ImportError as e:
    report.append(f"BLOCKER xlsx: {e}")
else:
    xlsx_dir = os.path.join(OUT, "xlsx"); os.makedirs(xlsx_dir, exist_ok=True)
    xlsx_path = os.path.join(xlsx_dir, f"{US}-testbook.xlsx")
    wb = Workbook()
    ws = wb.active; ws.title = "Scenarios"
    ws.append(["ID", "Title", "AC", "Condition", "Technique", "Priority",
               "Negative?", "Confidence", "Gherkin text"])
    n = 0
    for sc in scenarios:
        for rid, rtitle, steps in rows(sc):
            ws.append([rid, rtitle, sc["ac"], sc["condition"], sc["technique"], sc["priority"],
                       sc["negative"], sc["confidence"], "\n".join(background + steps)])
            n += 1
    ws2 = wb.create_sheet("Coverage matrix")
    for ln in read(MATRIX).splitlines():
        if ln.strip().startswith("|") and not re.fullmatch(r"\|[\s\-|]+\|", ln.strip()):
            ws2.append([c.strip() for c in ln.strip().strip("|").split("|")])
    ws3 = wb.create_sheet("Decisions & assumptions")
    ws3.append(["ID", "Type", "Statement", "Source checkpoint"])
    qa = read(os.path.join(SRC, "state", "02-understanding.md")).split("## Q&A log", 1)[1]
    for ln in qa.split("## Journey", 1)[0].splitlines():
        if ln.strip().startswith("| Q"):
            c = [x.strip() for x in ln.strip().strip("|").split("|")]
            kind = "open" if "[open]" in c[2] else "assumption"
            ws3.append([c[0], kind, f"{c[1]} -> {c[3]}", "state/02-understanding.md"])
    for wid, kind, stmt, srcf in [
        ("AC2-C1", "waiver", "P3 scope deferral, not generated (standing cited waiver)", "state/04-priorities.md"),
        ("AC3-C2", "waiver", "P3 scope deferral, not generated (standing cited waiver)", "state/04-priorities.md"),
        ("AC3-C3", "waiver", "P3 scope deferral, not generated (standing cited waiver)", "state/04-priorities.md"),
        ("3c-sort-filter", "waiver", "sort/filter coverage expansion not triggered, no such control observed", "state/03-design.md"),
        ("3c-pagination", "waiver", "pagination not triggered, small unpaginated table observed", "state/03-design.md"),
        ("3d-knowledge", "waiver", "no knowledge/index.md exists; degraded mode per shared rule 8", "state/03-design.md"),
        ("04-priorities", "simulated", "all priority scores accepted as proposed, non-interactive run", "state/04-priorities.md"),
    ]:
        ws3.append([wid, kind, stmt, srcf])
    wb.save(xlsx_path)
    from openpyxl import load_workbook
    chk = load_workbook(xlsx_path)
    report.append("re-parse OK  xlsx      sheets=" + ",".join(chk.sheetnames) +
                  f" scenarioRows={n} -> {xlsx_path}")

report.append(f"background steps flattened into every row: {len(background)} "
              f"({'none in this book' if not background else background})")
report.append(f"outlines exploded: {sum(1 for s in scenarios if s['outline'])}")
report.append(f"TestRail rows with an EMPTY Steps field (connector caveat): {empty_steps or 'none'}")
print("\n".join(report))
