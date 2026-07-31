#!/usr/bin/env python3
"""testbook-export skill, executed for real on US-EVAL-009 (skill-coverage wave 2026-07-30).

Implements plugins/qaia-core/skills/testbook-export/SKILL.md deliverables 1/3/4:
 - .feature copied as-is (deliverable 1, source of truth),
 - XLSX workbook, 3 sheets (deliverable 3),
 - Xray CSV per connectors/xray.md mapping table,
 - TestRail CSV per connectors/testrail.md mapping table.
Deliverable 2 (synthesis.md) already exists in the test book and is re-projected by copy.

Per connectors/*.md step 4: "materialize and run a short one-off script to do the mechanical
CSV writing + re-parse check (D42)". The re-parse check is at the bottom of this file.

Usage: python export_testbook.py <campaign-dir> <export-root>
"""
import csv, json, os, re, shutil, sys

PRIORITY_XRAY = {"P1": "Highest", "P2": "Medium", "P3": "Low"}
PRIORITY_TESTRAIL = {"P1": "Critical", "P2": "Medium", "P3": "Low"}
TECHNIQUE_TAGS = {"ep", "boundary", "decision-table", "state-transition", "use-case"}


def parse_feature(path):
    """Return (feature_title, background_steps, [scenario dicts])."""
    lines = open(path, encoding="utf-8").read().splitlines()
    feature_title, background, scenarios = None, [], []
    tags, comments, cur = [], [], None
    for raw in lines:
        s = raw.strip()
        if s.startswith("Feature:"):
            feature_title = s.split(":", 1)[1].strip()
        elif s.startswith("Background:"):
            cur = {"kind": "background", "steps": []}
        elif s.startswith("@"):
            if cur and cur.get("kind") == "scenario":
                scenarios.append(cur); cur = None
            tags = s.split()
            comments = []
        elif s.startswith("#"):
            comments.append(s.lstrip("# ").rstrip())
        elif s.startswith("Scenario Outline:") or s.startswith("Scenario:"):
            if cur and cur.get("kind") == "scenario":
                scenarios.append(cur)
            cur = {"kind": "scenario", "outline": s.startswith("Scenario Outline:"),
                   "title": s.split(":", 1)[1].strip(), "tags": tags, "comments": comments,
                   "steps": [], "examples": []}
            tags, comments = [], []
        elif s.startswith("Examples:") and cur:
            cur["in_examples"] = True
        elif re.match(r"^(Given|When|Then|And|But)\b", s):
            (background["steps"] if isinstance(background, dict) else
             (cur["steps"] if cur and cur["kind"] == "scenario" else []))
            if cur and cur.get("kind") == "background":
                cur["steps"].append(s)
            elif cur and cur["kind"] == "scenario":
                cur["steps"].append(s)
        elif s.startswith("|") and cur and cur.get("in_examples"):
            cur["examples"].append([c.strip() for c in s.strip("|").split("|")])
        elif s == "" and cur and cur.get("kind") == "background":
            background = cur["steps"]; cur = None
    if cur and cur.get("kind") == "scenario":
        scenarios.append(cur)
    if isinstance(background, dict):
        background = background["steps"]
    return feature_title, background, scenarios


def rows(scenarios):
    """Explode Scenario Outlines into one row per Examples line, ID suffixed -eN."""
    out = []
    for sc in scenarios:
        sid = next((t[1:] for t in sc["tags"] if t.startswith("@QAIA-")), "")
        pri = next((t[1:] for t in sc["tags"] if t[1:] in ("P1", "P2", "P3")), "")
        tech = [t[1:] for t in sc["tags"] if t[1:] in TECHNIQUE_TAGS]
        ac = next((t[1:] for t in sc["tags"] if re.fullmatch(r"@AC\d+", t)), "")
        cond = next((c.split(":", 1)[1].strip() for c in sc["comments"]
                     if c.startswith("condition:")), "")
        low = "@low-confidence" in sc["tags"]
        rationale = " ".join(c for c in sc["comments"] if not c.startswith("condition:"))
        ex = sc["examples"]
        if sc["outline"] and len(ex) > 1:
            for n, _ in enumerate(ex[1:], start=1):
                out.append(dict(sc=sc, id=f"{sid}-e{n}", title=f"{sc['title']} - Examples {n}",
                                pri=pri, tech=tech, ac=ac, cond=cond, low=low,
                                rationale=rationale, negative="@negative" in sc["tags"]))
        else:
            out.append(dict(sc=sc, id=sid, title=sc["title"], pri=pri, tech=tech, ac=ac,
                            cond=cond, low=low, rationale=rationale,
                            negative="@negative" in sc["tags"]))
    return out


def description(r):
    d = f"condition: {r['cond']}." if r["cond"] else ""
    if r["low"]:
        d += " low confidence — see open question in 02-understanding.md."
    if r["rationale"]:
        d += " " + r["rationale"]
    return d.strip()


def write_xray(rs, background, path):
    header = ["Issue Type", "Test Type", "Cucumber Test Type", "Summary",
              "Cucumber Scenario", "Labels", "Priority", "Description"]
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for r in rs:
            steps = background + r["sc"]["steps"]        # Background flattened in
            labels = " ".join([r["id"]] + ([r["pri"]] if r["pri"] else []) + r["tech"])
            w.writerow(["Test", "Automated[Cucumber]",
                        "Scenario Outline" if r["sc"]["outline"] else "Scenario",
                        r["title"], "\n".join(steps), labels,
                        PRIORITY_XRAY.get(r["pri"], ""), description(r)])
    return header


def split_gwt(steps):
    pre, act, exp, mode = [], [], [], "pre"
    for s in steps:
        kw = s.split()[0]
        if kw == "When":
            mode = "act"
        elif kw == "Then":
            mode = "exp"
        {"pre": pre, "act": act, "exp": exp}[mode].append(s)
    return pre, act, exp


def write_testrail(rs, background, feature_title, path):
    header = ["Title", "Type", "Priority", "Preconditions", "Steps",
              "Expected Result", "Section", "References"]
    empty_steps = []
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for r in rs:
            pre, act, exp = split_gwt(background + r["sc"]["steps"])
            if not act:
                empty_steps.append(r["id"])
            refs = " | ".join(filter(None, [r["id"], r["pri"], " ".join(r["tech"]), description(r)]))
            w.writerow([r["title"], "Automated", PRIORITY_TESTRAIL.get(r["pri"], ""),
                        "\n".join(pre), "\n".join(act), "\n".join(exp), feature_title, refs])
    return header, empty_steps


def write_xlsx(rs, background, base, matrix_md, understanding_md, priorities_md, path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Scenarios"
    ws.append(["ID", "Title", "AC", "Condition", "Technique", "Priority", "Negative?",
               "Confidence", "Gherkin text"])
    for r in rs:
        ws.append([r["id"], r["title"], r["ac"], r["cond"], " ".join(r["tech"]), r["pri"],
                   "yes" if r["negative"] else "no", "low" if r["low"] else "full",
                   "\n".join(background + r["sc"]["steps"])])
    ws2 = wb.create_sheet("Coverage matrix")
    for line in matrix_md.splitlines():
        if line.startswith("|") and not re.match(r"^\|[-| ]+\|$", line):
            ws2.append([c.strip() for c in line.strip("|").split("|")])
        elif line.strip():
            ws2.append([line.strip()])
    ws3 = wb.create_sheet("Decisions & assumptions")
    ws3.append(["ID", "Type", "Statement", "Source checkpoint"])
    for qid, rest in re.findall(r"^\| (Q\d+) \|(.*)$", understanding_md, re.M):
        cells = [c.strip() for c in rest.split("|")]
        kind = ("open" if "`[open]`" in rest else
                "assumption" if "`[assumption]`" in rest else "answered")
        ws3.append([qid, kind, cells[0] + " -> " + (cells[2] if len(cells) > 2 else ""),
                    "state/02-understanding.md"])
    # waivers / scope decisions from 03-design + 04-priorities
    for cond in ("AC2-C1", "AC3-C2", "AC3-C3"):
        ws3.append([cond, "waiver", "deferred to P3 by the default P1+P2 scope; not generated",
                    "state/04-priorities.md"])
    ws3.append(["3d-knowledge", "waiver",
                "no knowledge/index.md exists for this directory; degraded mode, nothing invented",
                "state/03-design.md"])
    for cond in ("sort/filter", "pagination", "CRUD-update", "role-axis", "auth-recovery"):
        ws3.append([cond, "waiver", "sub-step 3c bullet explicitly not triggered, reason stated",
                    "state/03-design.md"])
    ws3.append(["run", "simulated", "non-interactive campaign run: every VALIDATION gate simulated "
                "accepted-as-is; the pre-automation human Go/No-Go is pending-validation",
                "state/journey.md"])
    wb.save(path)
    return [ws.max_row - 1, ws2.max_row, ws3.max_row - 1]


def main(base, root):
    feat_src = os.path.join(base, "testbooks", "octoperf-petstore-cart.feature")
    title, background, scenarios = parse_feature(feat_src)
    rs = rows(scenarios)
    print(f"parsed: feature={title!r} background_steps={len(background)} "
          f"scenarios={len(scenarios)} exported_rows={len(rs)}")
    for r in rs:
        print(f"  {r['id']:24s} {r['pri']:3s} {r['ac']:4s} {r['cond']:8s} "
              f"steps={len(r['sc']['steps'])} outline={r['sc']['outline']}")

    os.makedirs(root, exist_ok=True)
    os.makedirs(os.path.join(root, "xray"), exist_ok=True)
    os.makedirs(os.path.join(root, "testrail"), exist_ok=True)
    shutil.copy(feat_src, os.path.join(root, "octoperf-petstore-cart.feature"))
    shutil.copy(os.path.join(base, "testbooks", "synthesis.md"), os.path.join(root, "synthesis.md"))

    xray_path = os.path.join(root, "xray", "US-EVAL-009-xray-import.csv")
    tr_path = os.path.join(root, "testrail", "US-EVAL-009-testrail-import.csv")
    xlsx_path = os.path.join(root, "US-EVAL-009-testbook.xlsx")

    hx = write_xray(rs, background, xray_path)
    ht, empty_steps = write_testrail(rs, background, title, tr_path)
    matrix = open(os.path.join(base, "testbooks", "coverage-matrix.md"), encoding="utf-8").read()
    und = open(os.path.join(base, "state", "02-understanding.md"), encoding="utf-8").read()
    pri = open(os.path.join(base, "state", "04-priorities.md"), encoding="utf-8").read()
    counts = write_xlsx(rs, background, base, matrix, und, pri, xlsx_path)

    print("\n--- re-parse check (D42) ---")
    for p, h in ((xray_path, hx), (tr_path, ht)):
        with open(p, encoding="utf-8", newline="") as f:
            rd = list(csv.reader(f))
        ok = rd[0] == h and len(rd) - 1 == len(rs)
        print(f"{'OK  ' if ok else 'FAIL'} {os.path.basename(p)}: header={rd[0] == h} "
              f"data_rows={len(rd) - 1} (expected {len(rs)})")
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    print(f"OK   {os.path.basename(xlsx_path)}: sheets={wb.sheetnames} "
          f"rows(scenarios/matrix/decisions)={counts}")
    print(f"\nempty TestRail Steps field (documented caveat) for: {empty_steps or 'none'}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
